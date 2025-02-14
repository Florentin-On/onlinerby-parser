import json
import logging
import os
import sys

import wx
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

APPDATA_PATH = os.path.join(os.getenv('APPDATA', os.getcwd()), 'OnlinerParser_Data')

MAIN_TITLE = 'Onliner Parser v1.0.0'

ABOUT_URL = 'https://github.com/Florentin-On'

MIN_SIZE = (1024, 768)
MAX_SIZE = (1024, 768)
MAIN_SIZE = (1024, 600)
LOG_SIZE = (1024, 68)

SETTINGS = {
    'panel_selection_default': 'Save on close',
}

CONTEXT_MENU_ITEMS = {
    1: 'copy_data',
    2: 'paste_data',
    3: 'copy_command_to_buffer',
}

heading_font = (20, wx.SWISS, wx.NORMAL, wx.BOLD, 'LucidaGrande')
small_heading_font = (14, wx.SWISS, wx.NORMAL, wx.BOLD, 'LucidaGrande')
default_font = (12, wx.SWISS, wx.NORMAL, wx.BOLD, 'LucidaGrande')

heading_excel_style = NamedStyle(name='Heading Style',
                                 font=Font(name="Arial Narrow", bold=True, size=14),
                                 alignment=Alignment(textRotation=90, horizontal='center'))
heading_simple_excel_style = NamedStyle(name='Heading Text Style',
                                        font=Font(name='Arial Narrow'),
                                        alignment=Alignment(textRotation=90, horizontal='center'), )
text_excel_style = NamedStyle(name='Text Style',
                              font=Font(name='Arial Narrow'),
                              alignment=Alignment(vertical='center'))
link_excel_style = NamedStyle(name='Link Style',
                              font=Font(name='Arial Narrow', color='000000FF', underline='single'),
                              alignment=Alignment(vertical='center'))
bool_true_excel_style = NamedStyle(name='Bool True Style',
                                   font=Font(name="Arial Narrow", color='375623'),
                                   fill=PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),
                                   alignment=Alignment(vertical='center'))
bool_false_excel_style = NamedStyle(name='Bool False Style',
                                    font=Font(name="Arial Narrow", color='595959'),
                                    fill=PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
                                    alignment=Alignment(vertical='center'))


def get_filters_parameters() -> dict:
    return {
        'general': _get_sub_filters_parameters(),
        'additional': _get_sub_filters_parameters(),
    }


def _get_sub_filters_parameters() -> dict:
    return {
        'parameters_dict': {},
        'parameters_dict_from': {},
        'parameters_dict_to': {},
        'parameters_number_range_from': {},
        'parameters_number_range_to': {},
        'parameters_checkbox': {},
    }


def get_main_parameters() -> dict:
    return {x: True for x in
            ['Картинка', 'Бренд', 'Модель и ссылка на Onliner', 'Тип', 'Цена минимальная', 'Цена максимальная',
             'Количество предложений', 'Оценка и Количество отзывов', 'Стикеры']}


def create_font(data):
    """"""
    size, family, style, weight, face = data
    font = wx.Font(size, family, style, weight)
    return font


def safe_load_json(path_to_json: str) -> dict:
    try:
        return json.load(open(path_to_json, 'r'))
    except Exception as error:
        logging.warning(f'Can\'t load config: {path_to_json}. Error: {error}')
        return {}


def resource_path(relative_path: str) -> str:
    """
    The function creates a path depending on the code used. Source code - the current working directory is used,
    determined by the presence of the start_source file. There is no start_source in the build,
    which means sys._MEIPASS will be used.
    This variable is created by the build itself after launch and stores unpacked resources along this path.
    :param relative_path: relative path to a file
    :type relative_path: str
    :return: return absolute path to file
    :rtype: str or None
    """
    base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.getcwd()
    logging.info(f'Base path: {base_path}')
    return os.path.join(base_path, relative_path)
