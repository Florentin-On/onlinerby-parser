import io
import logging
import os
import time
import traceback
from copy import deepcopy
from threading import Thread
from typing import Optional, Dict, Any, Union

import openpyxl
import wx
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

from app.common.constants import get_filters_parameters, heading_excel_style, \
    heading_simple_excel_style, text_excel_style, link_excel_style, bool_true_excel_style, bool_false_excel_style, \
    get_main_parameters, heading_font, create_font
from app.common.safe_requesters import safe_get_requester
from app.common_ui.dialogs import dialog, confirmation_dialog, confirmation_with_cancel_dialog
from app.multiparse.multiparse_dialogs import TemplateMultiparseDialog


class Multiparse(wx.Panel):
    def __init__(self, parent, size):
        wx.Panel.__init__(self, parent=parent, size=size)
        self.parent = parent

        # Флаг для того, чтобы сбрасывать список параметров, выбранных для выгрузки в отчет, в случае изменения фильтра
        self.filterSpecified = False
        # Словарь вида "Категория -> Группа -> Раздел"
        self.categories = self.load_categories()
        # Словарь с уже подгруженными параметрами Разделов. Ключ: название Раздела.
        # Значение: кортеж URL раздела (facets) и данные по нему
        self.sections_parameters = {}
        # Параметры фильтров, которые разделены на "Основные" и "Дополнительные"
        self.filters_parameters = get_filters_parameters()
        # Постоянные параметры, которые есть у каждого товара. Используется для задания перечня того, что выводить в
        # отчет из основных параметров. Ключ: название параметра. Значение: булево значение
        self.main_product_parameters = get_main_parameters()
        # Данные по первому продукту для указанного фильтра. Ключ: URL с фильтром. Значение: параметры продукта в виде
        # словаря. Словарь с параметрами: ключ - название группы параметров, значение - список параметров группы
        self.product_parameters = {}
        # Выбранные параметры продуктов, которые необходимо выводить в отчет от выбранного Раздела продуктов
        # с примененным фильтром. Ключ: название группы параметров. Значение: список параметров группы в виде словаря
        # с их значениями, где ключ - параметр, значение - булево значение параметра.
        self.filtered_product_parameters = {}

        self.SetFont(create_font(heading_font))
        main_sizer = wx.BoxSizer(orient=wx.HORIZONTAL)
        left_sizer = wx.BoxSizer(orient=wx.VERTICAL)
        self.highest_categories_label = wx.StaticText(self, label='Раздел')
        self.product_category_combobox = wx.ComboBox(self, -1, name='product_category_combobox')
        left_sizer_line_1 = wx.BoxSizer(orient=wx.VERTICAL)
        left_sizer_line_1_1 = wx.BoxSizer(orient=wx.VERTICAL)
        left_sizer_line_1.Add(self.highest_categories_label, flag=wx.ALIGN_CENTER | wx.UP, border=10)
        left_sizer_line_1_1.Add(self.product_category_combobox, flag=wx.ALL | wx.EXPAND, border=5)

        left_sizer_line_2 = wx.BoxSizer(orient=wx.VERTICAL)
        left_sizer_line_2_1 = wx.BoxSizer(orient=wx.VERTICAL)
        self.categories_label = wx.StaticText(self, label='Категория')
        self.product_group_combobox = wx.ComboBox(self, -1, name='product_group_combobox')
        left_sizer_line_2.Add(self.categories_label, flag=wx.ALIGN_CENTER | wx.UP, border=10)
        left_sizer_line_2_1.Add(self.product_group_combobox, flag=wx.ALL | wx.EXPAND, border=5)

        left_sizer_line_3 = wx.BoxSizer(orient=wx.VERTICAL)
        left_sizer_line_3_1 = wx.BoxSizer(orient=wx.VERTICAL)
        self.sub_categories_label = wx.StaticText(self, label='Податегория')
        self.product_section_combobox = wx.ComboBox(self, -1, name='product_section_combobox')
        left_sizer_line_3.Add(self.sub_categories_label, flag=wx.ALIGN_CENTER | wx.UP, border=10)
        left_sizer_line_3_1.Add(self.product_section_combobox, flag=wx.ALL | wx.EXPAND, border=5)

        left_sizer_line_4 = wx.BoxSizer(orient=wx.VERTICAL)
        self.select_search_general_params_button = wx.Button(self, label='Задать основные параметры фильтра',
                                                             name='general')
        self.select_search_general_params_button.Disable()
        left_sizer_line_4.Add(self.select_search_general_params_button, flag=wx.ALL | wx.ALIGN_CENTER, border=10)

        left_sizer_line_5 = wx.BoxSizer(orient=wx.VERTICAL)
        self.select_search_add_params_button = wx.Button(self, label='Задать дополнительные параметры фильтра',
                                                         name='additional')
        self.select_search_add_params_button.Disable()
        left_sizer_line_5.Add(self.select_search_add_params_button,
                              flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER, border=10)

        left_sizer_line_6 = wx.BoxSizer(orient=wx.VERTICAL)
        self.select_needed_params_button = wx.Button(self, label='Выбрать параметры для формирования отчета',
                                                     name='select_needed_params_button')
        self.select_needed_params_button.Disable()
        left_sizer_line_6.Add(self.select_needed_params_button,
                              flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER, border=10)

        left_sizer_line_7 = wx.BoxSizer(orient=wx.VERTICAL)
        self.generate_report_button = wx.Button(self, label='Сформировать отчет с заданными параметрами',
                                                name='generate_report_button')
        self.generate_report_button.Disable()
        left_sizer_line_7.Add(self.generate_report_button,
                              flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER, border=10)

        left_sizer.Add((0, 0), proportion=1, flag=wx.ALL | wx.EXPAND, border=5)
        left_sizer.Add(left_sizer_line_1, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_1_1, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_2, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_2_1, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_3, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_3_1, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_4, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_5, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_6, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add(left_sizer_line_7, proportion=1, flag=wx.ALL | wx.EXPAND, border=0)
        left_sizer.Add((0, 0), proportion=1, flag=wx.ALL | wx.EXPAND, border=5)

        main_sizer.Add(left_sizer, proportion=1, flag=wx.LEFT | wx.RIGHT | wx.EXPAND, border=100)
        self.SetSizer(main_sizer)

        self.update_categories(self.product_category_combobox)

        self.product_category_combobox.Bind(wx.EVT_TEXT, self.categories_changes)
        self.product_group_combobox.Bind(wx.EVT_TEXT, self.categories_changes)
        self.product_section_combobox.Bind(wx.EVT_TEXT, self.categories_changes)

        self.select_search_general_params_button.Bind(wx.EVT_BUTTON, self.open_search_params)
        self.select_search_add_params_button.Bind(wx.EVT_BUTTON, self.open_search_params)
        self.select_needed_params_button.Bind(wx.EVT_BUTTON, self.open_search_params)
        self.generate_report_button.Bind(wx.EVT_BUTTON, self.generate_report)

    @staticmethod
    def load_categories() -> dict:
        """
        Загрузка всех категорий из каталога Онлайнера и формирование трехуровневого словаря вида
        Категория -> Группа -> Раздел
        """
        category_dict: dict = {}
        categories_info = safe_get_requester('https://catalog.api.onliner.by/navigation/elements', [])
        for category in categories_info:
            category_title = category.get('title')
            if category_title and category_title not in category_dict and category.get('slug').lower() != 'prime':
                category_dict[category_title]: dict = {}
                category_groups_info = safe_get_requester(category.get('groups_url', ''), [])
                for group in category_groups_info:
                    group_title = group.get('title')
                    if group_title and group_title not in category_dict[category_title]:
                        category_dict[category_title][group_title]: dict = {}
                        group_sections_info = group.get('links', [])
                        for section in group_sections_info:
                            section_title = section.get('title')
                            if section_title and section_title not in category_dict[category_title][group_title]:
                                category_dict[category_title][group_title][section_title]: dict = section.get(
                                    'source_urls', {})
        return category_dict

    def update_categories(self, product_combobox: wx.ComboBox) -> None:
        """
        Обновление ComboBox в зависимости от того, кто вызвал данный метод
        :param product_combobox: wx.ComboBox, вызывающий метод
        """
        categories_keys = []
        combobox_name = product_combobox.GetName()
        if combobox_name == 'product_category_combobox':
            categories_keys += list(self.categories)
        elif combobox_name == 'product_group_combobox':
            categories_keys += list(self.categories[self.product_category_combobox.GetValue()])
        elif combobox_name == 'product_section_combobox':
            categories_keys += list(self.categories[self.product_category_combobox.GetValue()][
                                        self.product_group_combobox.GetValue()])
        product_combobox.Append(categories_keys)

    def categories_changes(self, event: wx.Event) -> None:
        """
        Метод смены выбора категории товаров. При выборе Раздела - подгружает информацию по его параметрам
        :param event: wx.Event
        """
        combobox_value = event.GetEventObject().GetValue()
        if combobox_value == '':  # Этот же метод вызывается при вызове Clear() у wx.ComboBox
            return None
        key = event.GetEventObject().GetName()

        if key == 'product_category_combobox':
            self.product_group_combobox.Clear()
            self.product_section_combobox.Clear()

            self.update_categories(self.product_group_combobox)
        elif key == 'product_group_combobox':
            self.product_section_combobox.Clear()

            self.update_categories(self.product_section_combobox)
        elif key == 'product_section_combobox':
            category = self.product_category_combobox.GetValue()
            group = self.product_group_combobox.GetValue()
            section_url = self.categories[category][group][combobox_value]['catalog.schema.facets']
            if combobox_value not in self.sections_parameters:
                self.sections_parameters[combobox_value] = (section_url, safe_get_requester(section_url, []))
            self.select_search_general_params_button.Enable()
            self.select_search_add_params_button.Enable()
            self.select_needed_params_button.Enable()
            self.generate_report_button.Enable()

            # Сбрасываем заданные фильтры и перечень выводимых в отчет параметров
            self.filters_parameters = get_filters_parameters()
            self.filtered_product_parameters = {}

        if self.product_section_combobox.GetValue() == '':
            self.select_search_general_params_button.Disable()
            self.select_search_add_params_button.Disable()
            self.select_needed_params_button.Disable()
            self.generate_report_button.Disable()

    def open_search_params(self, event: wx.Event) -> None:
        """
        Метод открытия окна фильтров
        :param event: wx.Event
        """
        section = self.product_section_combobox.GetValue()

        filter_group_name = event.GetEventObject().GetName()
        if self.product_section_combobox.GetValue() != '':
            dlg = (
                TemplateMultiparseDialog(self.sections_parameters[section], filter_group_name, self)
                if filter_group_name in ('general', 'additional')
                else self.set_report_parameters_dialog(filter_group_name)
            )
            if dlg is not None:
                dlg.Center()
                dlg.ShowModal()
                dlg.Destroy()

    def set_report_parameters_dialog(self, filter_group_name: str) -> Optional[TemplateMultiparseDialog]:
        """
        Метод подгружает данные по первому товару, для которого применимы текущие выбранные параметры фильтра. Затем
        открывает окно с выбором параметров для формирования отчета на основании полученных параметров из товара.
        :param filter_group_name: Название кнопки, которое передается в открывающееся окно выбора параметров
        """
        link = self.get_link_from_filters()
        if link not in self.product_parameters:
            try:
                with wx.BusyInfo('Подгружаем параметры товаров...'):
                    products_with_filter = safe_get_requester(link.lower(), {})
                if products_with_filter.get('products'):
                    self.product_parameters[link] = self.get_all_product_parameters(
                        products_with_filter['products'][0]['url'])
                else:
                    dialog('Не найдено', 'Товары с заданным фильтром не найдены!')
                    return None
            except Exception as err:
                msg = 'Ошибка:\n' + str(err) + '\n\n' + 'Подробности в логах программы.'
                traceback.print_exc()
                dialog(caption='Ошибка', message=msg, style=wx.ICON_ERROR)
                return None
        # Если перед открытием диалога выбора выводимых в отчет параметров поменялись параметры фильтра - нужно сбросить
        # текущий выбор параметров для вывода в отчет
        if self.filterSpecified:
            self.filtered_product_parameters = {}
            self.filterSpecified = False
        return TemplateMultiparseDialog(self.product_parameters[link], filter_group_name, self)

    def get_link_from_filters(self) -> str:
        """
        Метод, который возвращает URL поиска товаров с заданными текущими фильтрами
        """
        products_link = \
            self.categories[self.product_category_combobox.GetValue()][self.product_group_combobox.GetValue()][
                self.product_section_combobox.GetValue()]['catalog.schema.products']
        products_link += '&' if '?' in products_link else '?'
        filter_parts = []

        def process_dict_parameters(value, key_format):
            for param_ids, param_values in value.items():
                if param_values:
                    all_ids = [
                        str(i['id'])
                        for i in
                        self.sections_parameters[self.product_section_combobox.GetValue()][1]['dictionaries'][param_ids]
                        if i['name'] in param_values
                    ]
                    filter_parts.extend([key_format.format(name=param_ids, id=num, value=p_value) for num, p_value in
                                         enumerate(all_ids)])

        for filter_type in self.filters_parameters.values():
            for parameter_id, value in filter_type.items():
                if value:
                    if parameter_id == 'parameters_dict':
                        process_dict_parameters(value, '{name}[{id}]={value}')
                    elif parameter_id == 'parameters_dict_from':
                        process_dict_parameters(value, '{name}[from]={value}')
                    elif parameter_id == 'parameters_dict_to':
                        process_dict_parameters(value, '{name}[to]={value}')
                    elif parameter_id in ('parameters_number_range_from', 'parameters_number_range_to'):
                        for param_id, param_value in value.items():
                            if param_value:
                                filter_parts.append(f'{param_id}[{parameter_id.split("_")[-1]}]={param_value}')
                    elif parameter_id == 'parameters_checkbox':
                        for param_id, param_value in value.items():
                            if param_value:
                                filter_parts.append(f'{param_id}={int(param_value)}')
        return products_link + '&'.join(filter_parts)

    @staticmethod
    def get_all_product_parameters(product_url: str, only_headers: bool = True) -> Dict[str, Union[list, dict]]:
        """
        Метод, возвращающий либо только названия всех имеющихся параметров указанного продукта,
        либо и названия, и значения этих параметров
        :param product_url: URL продукта
        :param only_headers: Флаг, отвечающий за возврат только списка параметров или вместе с их значениями
        """
        product_info = safe_get_requester(product_url + '?include=parameters', {})
        return_parameters = {}
        product_parameters = product_info.get('parameters', [])
        if product_parameters:
            for parameter_group in product_parameters:
                product_group_parameters = parameter_group.get('parameters')
                return_parameters[parameter_group['name']] = (
                    [param['name'] for param in product_group_parameters] if only_headers
                    else {param['name']: param['value'] for param in product_group_parameters}
                )
        return return_parameters

    def generate_report(self, event) -> None:
        """
        Метод создания отчета по выбранному фильтру и с заданными параметрами для вывода в файл отчета
        """
        link = self.get_link_from_filters()
        filtered_products_dict = safe_get_requester(link.lower(), {})
        pages_count = filtered_products_dict['page']['last']
        total_product_count = filtered_products_dict['total']
        if not total_product_count:
            dialog('Не найдено', 'Товары с указанным фильтром не найдены!')
            return None
        result = confirmation_with_cancel_dialog('Основные параметры', 'Выгрузить в отчет только основные параметры?')
        if result in (wx.YES, wx.NO):
            only_main_parameters = result == wx.YES
        else:
            return None
        with wx.FileDialog(self, "Выберите место и имя для сохранения отчета",
                           wildcard="Excel files (*.xlsx)|*.xlsx|All files (*.*)|*.*",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return None

            new_file = True
            pathname = fileDialog.GetPath()
            try:
                if os.path.isfile(pathname):
                    wb = openpyxl.load_workbook(pathname)
                    if 'DEV_ONLINER_PARSER' in wb.sheetnames:
                        stored_link = wb['DEV_ONLINER_PARSER']['A2'].value
                        if stored_link == link:
                            # TODO: подумать над тем, что в тот же файл можно начать выгружать с другими выбранными параметрами
                            if confirmation_dialog("Внимание", "Файл, который вы выбрали, уже содержит товары с "
                                                               "заданным фильтром. Продолжить выгрузку в данный файл?") == wx.YES:
                                new_file = False
                            else:
                                return None
                        elif confirmation_dialog("Внимание",
                                                 "Файл, который вы выбрали, уже содержит товары другого "
                                                 "фильтра. Вы утеряете все данные из данного файла. Продолжить?") == wx.NO:
                            return None
                    elif confirmation_dialog("Внимание", "Вы утеряете все данные из выбранного файла. "
                                                         "Продолжить?") == wx.NO:
                        return None
                selected_main_parameters = {param: value for param, value in self.main_product_parameters.items()}
                selected_add_parameters = {}
                if not only_main_parameters:
                    selected_add_parameters = self.get_parameters_for_workbook(link, filtered_products_dict)
                if new_file:
                    with wx.BusyInfo("Создается файл отчета..."):
                        self.create_empty_excel_table(pathname, selected_add_parameters, link, only_main_parameters)
                wb = openpyxl.load_workbook(pathname)
                goods_amount = wb['DEV_ONLINER_PARSER']['A1'].value
                progress_window = wx.GenericProgressDialog('Товары выгружаются', 'Прогресс\nВыгружено {} из {}'
                                                           .format('0', str(total_product_count - goods_amount)),
                                                           maximum=total_product_count - goods_amount,
                                                           parent=self.parent,
                                                           style=wx.PD_APP_MODAL | wx.PD_ELAPSED_TIME |
                                                                 wx.PD_REMAINING_TIME | wx.PD_ESTIMATED_TIME |
                                                                 wx.PD_AUTO_HIDE | wx.PD_SMOOTH | wx.PD_CAN_ABORT)

                Thread(target=self.process_report, daemon=True,
                       args=(pathname, link, filtered_products_dict, pages_count, selected_main_parameters,
                       selected_add_parameters, wb, progress_window),
                       kwargs={'only_main_parameters': only_main_parameters}).start()
            except Exception as err:
                msg = 'Ошибка создания отчета:\n' + str(err) + '\n\n' + 'Подробности в логах программы.'
                traceback.print_exc()
                dialog(caption='Ошибка', message=msg, style=wx.ICON_ERROR)
        event.Skip()

    def get_parameters_for_workbook(self, link: str, filtered_products_dict: dict) -> Dict[str, list]:
        """
        Метод, возвращающий список параметров, который будет выгружен в отчет
        :param link: URL списка продуктов с установленным фильтром
        :param filtered_products_dict: словарь с данными по продуктам по заданному в link фильтре
        """
        all_headings = {}
        if self.filtered_product_parameters:
            for params_group, param in self.filtered_product_parameters.items():
                for param_name, param_flag in param.items():
                    if param_flag:
                        if params_group not in all_headings:
                            all_headings[params_group] = []
                        all_headings[params_group].append(param_name)
        if not all_headings:
            if link not in self.product_parameters:
                self.product_parameters[link] = self.get_all_product_parameters(
                    filtered_products_dict['products'][0]['url'])
            for params_group, param in self.product_parameters[link].items():
                all_headings[params_group] = param
        return all_headings

    def create_empty_excel_table(self, pathname: str, selected_parameters: dict, link: str, only_main_parameters: bool):
        """
        Метод создает пустую таблицу со скрытым листом, в которую прописывается URL текущего фильтра для возможности
        прерывания и последующего возобновления выгрузки в один и тот же файл
        :param pathname: Путь до файла отчета
        :param selected_parameters: Все параметры, которые должны отображаться в отчете
        :param link: URL списка продуктов с установленным фильтром
        :param only_main_parameters: Флаг, отвечающий за необходимость выгрузки только основных параметров продуктов
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Report'
        id = 1
        wb.add_named_style(heading_excel_style)
        wb.add_named_style(text_excel_style)
        wb.add_named_style(link_excel_style)
        wb.add_named_style(bool_true_excel_style)
        wb.add_named_style(bool_false_excel_style)
        wb.add_named_style(heading_simple_excel_style)
        process_headings = [heading for heading in self.main_product_parameters if
                            self.main_product_parameters[heading]]
        for column_id, title in enumerate(process_headings, start=id):
            ws.cell(column=column_id, row=1, value=title).style = 'Heading Style'
            # ws.column_dimensions[openpyxl.utils.get_column_letter(id)].width = len(title)
            id = column_id + 1

        if not only_main_parameters:
            for group_heading in selected_parameters:
                ws.cell(column=id, row=1, value=group_heading).style = 'Heading Style'
                id += 1
                # ws.column_dimensions[openpyxl.utils.get_column_letter(id)].width = len(group)
                for heading in selected_parameters[group_heading]:
                    # ws.column_dimensions[openpyxl.utils.get_column_letter(id)].width = len(heading)
                    ws.cell(column=id, row=1, value=heading).style = 'Heading Text Style'
                    id += 1
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        ws = wb.create_sheet('DEV_ONLINER_PARSER')
        ws.sheet_state = 'hidden'
        ws['A1'] = 0
        ws['A2'] = link
        wb.save(pathname)
        wb.close()

    def process_report(self,
                       pathname: str,
                       search_link: str,
                       filtered_products_dict: dict,
                       pages_count: int,
                       selected_main_parameters: dict,
                       selected_add_parameters: dict,
                       wb: Workbook,
                       progress_window: wx.GenericProgressDialog,
                       only_main_parameters: bool = False) -> None:
        """
        Метод, который формирует отчет в Excel по заданному фильтру поиска и заданным параметрам для выгрузки
        :param pathname: Путь до Excel файла с отчетом
        :param search_link: URL со списком продуктов по заданным фильтрам
        :param filtered_products_dict: Словарь с первой страницей продуктов, отфильтрованными по заданным параметрам
        фильтров
        :param pages_count: Количество страниц продуктов по заданному фильтру
        :param selected_main_parameters: Выбранные для выгрузки в отчет основные параметры
        :param selected_add_parameters: Выбранные для выгрузки в отчет продукт-специфичные параметры
        :param wb: Рабочая книга Excel, в которую ведется запись
        :param progress_window: Окно отображения прогресса выгрузки
        :param only_main_parameters: Флаг, отвечающий за необходимость выгрузки только основных параметров (ускоряет
        выгрузку, т.к. не требует запроса информации по каждому продукту в отдельности)
        """
        goods_amount = wb['DEV_ONLINER_PARSER']['A1'].value
        current_page_products_dict = deepcopy(filtered_products_dict)
        start_index = 1
        if goods_amount != 0:
            start_index = goods_amount // 30 + 1
        sleep = 0.1
        progress_window_bar = 0
        delta_iterate_value = 2
        for i in range(start_index, pages_count + 1):
            break_flag = False
            if i != 1:
                proceeded_link = search_link
                time.sleep(sleep)
                proceeded_link += 'page=' + str(i)
                current_page_products_dict = safe_get_requester(proceeded_link.lower(), {})
            ws = wb.active
            start_page_index = goods_amount % 30
            for product in current_page_products_dict['products'][start_page_index:]:
                product_html = product['html_url']
                id = 1
                row_id = goods_amount + delta_iterate_value
                for main_header, flag in selected_main_parameters.items():
                    if flag:
                        ws.cell(column=id, row=row_id).style = 'Text Style'
                        if main_header == 'Картинка':
                            if product['images']['header']:
                                img_link = product['images']['header']
                                r = safe_get_requester(img_link, raw_response=True)
                                if r is not None:
                                    image_file = io.BytesIO(r.content)
                                    img = Image(image_file)
                                    img.anchor = ws.cell(column=1, row=row_id).coordinate
                                    img_width = img.width / 7.0
                                    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, img_width)
                                    ws.row_dimensions[row_id].height = img.height / 1.33
                                    ws.add_image(img)
                            else:
                                ws.cell(column=id, row=row_id, value='-')
                        elif main_header == 'Бренд':
                            brand = product['full_name'].replace(product['name'], '')
                            ws.cell(column=id, row=row_id, value=brand)
                        elif main_header == 'Модель и ссылка на Onliner':
                            ws.cell(column=id, row=row_id, value='=HYPERLINK("{}", "{}")'
                                    .format(product_html, product['name'])).style = 'Link Style'
                        elif main_header == 'Тип':
                            ws.cell(column=id, row=row_id, value=product['name_prefix'])
                        elif main_header == 'Цена минимальная':
                            ws.cell(column=id, row=row_id,
                                    value=float(product['prices']['price_min']['amount']) if product['prices'] else '-')
                        elif main_header == 'Цена максимальная':
                            ws.cell(column=id, row=row_id,
                                    value=float(product['prices']['price_max']['amount']) if product['prices'] else '-')
                        elif main_header == 'Количество предложений':
                            ws.cell(column=id, row=row_id,
                                    value=product['prices']['offers']['count'] if product['prices'] else '-')
                        elif main_header == 'Оценка и Количество отзывов':
                            ws.cell(column=id, row=row_id,
                                    value=f'{product["reviews"]["rating"] / 10} ({product["reviews"]["count"]})')
                        elif main_header == 'Стикеры':
                            stickers = product['stickers']
                            sticker_names = [sticker['label'] for sticker in stickers] if stickers else []
                            text = ', '.join(sticker_names) if sticker_names else '-'
                            ws.cell(column=id, row=row_id, value=text)
                        id += 1

                if not only_main_parameters:
                    time.sleep(sleep)
                    product_api_url = product['url']
                    selected_product_parameters = self.get_selected_product_parameters(product_api_url,
                                                                                       selected_add_parameters)
                    need_row_to_increase = 0
                    for group in selected_product_parameters.values():
                        ws.cell(column=id, row=row_id, value=' ').fill = \
                            PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        id += 1
                        for parameter_name, parameter_value_list in group.items():
                            """
                            if isinstance(parameter_value, list):
                                need_row_to_increase = len(parameter_value) - 1
                                if need_row_to_increase > 1:
                                    for merge_id in range(1, id + sum(
                                            [len(i.values()) + 1 for i in selected_product_parameters.values()])):
                                        if merge_id != id:
                                            ws.merge_cells(start_column=merge_id,
                                                           start_row=row_id,
                                                           end_column=merge_id,
                                                           end_row=row_id + need_row_to_increase)
                                for list_value in parameter_value:
                                    ws.cell(column=id, row=row_id + parameter_value.index(list_value),
                                            value='=HYPERLINK("{}", "{}")'
                                            .format(list_value[1], list_value[0])).style = 'Link Style'
                            """
                            if parameter_value_list == '-':
                                ws.cell(column=id, row=row_id, value=str(parameter_value_list)).style = 'Text Style'
                            elif len(parameter_value_list) == 1:
                                param_type = parameter_value_list[0]['type']
                                if param_type == 'link':
                                    param_links = parameter_value_list[0]['link']
                                    if isinstance(param_links, dict):
                                        ws.cell(column=id, row=row_id, value='=HYPERLINK("{}", "{}")'
                                                .format(param_links['source_urls']['catalog.product.web'],
                                                        param_links['title'])).style = 'Link Style'
                                    else:
                                        logging.warning(f'Too much links for parameter or it is not dict. '
                                                        f'Product: {product_api_url}. Parameter: '
                                                        f'{parameter_value_list[0]}')
                                elif param_type == 'bool':
                                    parameter_value = parameter_value_list[0]['value']
                                    ws.cell(column=id, row=row_id, value=str(parameter_value)).style = \
                                        'Bool True Style' if parameter_value else 'Bool False Style'
                                elif param_type == 'string':
                                    parameter_value = parameter_value_list[0]['value']
                                    ws.cell(column=id, row=row_id, value=str(parameter_value)).style = 'Text Style'
                                else:
                                    logging.warning(f'Unknown type of parameter {param_type}. '
                                                    f'Product: {product_api_url}. Parameter: {parameter_value_list[0]}')
                            elif len(parameter_value_list) == 2:
                                parameter_value = {'string': None, 'bool': None}
                                for parameter in parameter_value_list:
                                    param_type = parameter['type']
                                    if param_type in parameter_value:
                                        parameter_value[param_type] = parameter['value']
                                    else:
                                        logging.warning(f'Unknown type of parameter {param_type}. Product: '
                                                        f'{product_api_url}. Parameter(-s): {parameter_value_list[0]}')
                                ws.cell(column=id, row=row_id, value=str(parameter_value['string'])).style = \
                                    'Bool True Style' if parameter_value['bool'] else 'Bool False Style'
                            id += 1
                    delta_iterate_value += need_row_to_increase
                goods_amount += 1
                wb['DEV_ONLINER_PARSER']['A1'].value = goods_amount
                progress_window_bar += 1
                product_name_cutted = f'{product["full_name"][:25]}...'
                if (progress_window.WasCancelled() or
                        not progress_window.Update(progress_window_bar,
                                                   newmsg='Прогресс\nВыгружено {} из {}. Выгружаю продукт: {}'.format(
                                                       progress_window_bar, str(progress_window.GetRange()),
                                                       product_name_cutted))[0]):
                    break_flag = True
                    break
            if break_flag:
                break

        caption, message, icon = 'Завершено', 'Выгрузка успешно завершена!', wx.OK
        try:
            wb.save(pathname)
            if progress_window.WasCancelled():
                caption, message, icon = 'Прервано', 'Выгрузка прервана пользователем', wx.ICON_WARNING
        except PermissionError as err:
            msg = 'Ошибка создания отчета - указанный файл был открыт во время выгрузки:\n' + str(
                err) + '\n\nПодробности в логах программы.'
            traceback.print_exc()
            caption, message, icon = 'Ошибка', msg, wx.ICON_ERROR
        dialog(caption, message, icon)
        progress_window.Destroy()
        try:
            if os.path.isfile(pathname):
                os.startfile(pathname)
        except Exception as error:
            traceback.print_exc()
            logging.exception(error)
        return None

    def get_selected_product_parameters(self, product_link: str, selected_parameters: dict) -> Dict[
        str, Dict[str, list]]:
        """
        Метод, возвращающий все параметры выбранного продукта с учетом выбранных параметров для выгрузки в отчет
        :param product_link: URL выбранного продукта
        :param selected_parameters: Выбранные для выгрузки в отчет параметры
        """
        product_proceeded_parameters: Dict[str, Dict[str, Any]] = self.get_all_product_parameters(product_link,
                                                                                                  only_headers=False)

        return_parameters: Dict[str, Dict[str, Any]] = {}
        for sel_group, sel_group_list in selected_parameters.items():
            return_parameters[sel_group] = {}
            for sel_group_value in sel_group_list:
                return_parameters[sel_group][sel_group_value] = '-'
        for group_name, group_list in product_proceeded_parameters.items():
            if group_name in selected_parameters:
                for group_parameter in group_list:
                    if group_parameter in selected_parameters[group_name]:
                        return_parameters[group_name][group_parameter] = product_proceeded_parameters[group_name][
                            group_parameter]
        return return_parameters
